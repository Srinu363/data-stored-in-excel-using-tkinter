import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook

def save_data():
    name = name_entry.get()
    year = year_entry.get()
    roll_number = roll_number_entry.get()
    phone_number = phone_number_entry.get()

    if name and year and roll_number and phone_number:
        try:
            wb = load_workbook('student_data.xlsx')
            sheet = wb.active
        except FileNotFoundError:
            wb = Workbook()
            sheet = wb.active
            sheet.append(["Name", "Year", "Roll Number", "Phone Number"])

        sheet.append([name, year, roll_number, phone_number])
        wb.save('student_data.xlsx')

        messagebox.showinfo("Success", "Data successfully stored in Excel sheet.")
        name_entry.delete(0, tk.END)
        year_entry.delete(0, tk.END)
        roll_number_entry.delete(0, tk.END)
        phone_number_entry.delete(0, tk.END)
    else:
        messagebox.showerror("Error", "Please fill all the fields.")

app = tk.Tk()
app.title("Student Data Entry")
app.geometry("400x200")

name_label = tk.Label(app, text="Name:")
name_label.pack()
name_entry = tk.Entry(app)
name_entry.pack()

year_label = tk.Label(app, text="Year:")
year_label.pack()
year_entry = tk.Entry(app)
year_entry.pack()

roll_number_label = tk.Label(app, text="Roll Number:")
roll_number_label.pack()
roll_number_entry = tk.Entry(app)
roll_number_entry.pack()

phone_number_label = tk.Label(app, text="Phone Number:")
phone_number_label.pack()
phone_number_entry = tk.Entry(app)
phone_number_entry.pack()

save_button = tk.Button(app, text="Save Data", command=save_data)
save_button.pack()

app.mainloop()
